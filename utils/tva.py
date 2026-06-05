# -*- coding: utf-8 -*-
"""
Module Aide TVA CA3 / CA12
Calcul, vérification et aide à la déclaration TVA France
SMD Global Consulting LLC - DGFiP / PCG France
"""
import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io
from datetime import datetime

# ─────────────────────────────────────────────
# CONSTANTES TVA FRANCE
# ─────────────────────────────────────────────
TAUX_TVA = {
    "20% — Taux normal": 0.20,
    "10% — Taux intermédiaire": 0.10,
    "5,5% — Taux réduit": 0.055,
    "2,1% — Taux super-réduit": 0.021,
    "0% — Exonéré / Export": 0.0,
}

RUBRIQUES_CA3 = {
    "Ventes soumises à 20%": {"taux": 0.20, "ligne_ca3": "A1", "compte_pcg": "70x"},
    "Ventes soumises à 10%": {"taux": 0.10, "ligne_ca3": "A2", "compte_pcg": "70x"},
    "Ventes soumises à 5,5%": {"taux": 0.055, "ligne_ca3": "A3", "compte_pcg": "70x"},
    "Ventes soumises à 2,1%": {"taux": 0.021, "ligne_ca3": "A4", "compte_pcg": "70x"},
    "Acquisitions intracommunautaires (20%)": {"taux": 0.20, "ligne_ca3": "B1", "compte_pcg": "401/445"},
    "Autoliquidation achats (20%)": {"taux": 0.20, "ligne_ca3": "B2", "compte_pcg": "401/445"},
    "Ventes exonérées / exports": {"taux": 0.0, "ligne_ca3": "E1", "compte_pcg": "70x"},
}

COMPTES_TVA_PCG = {
    "44566 — TVA déductible sur ABS": "TVA déductible sur biens et services",
    "44562 — TVA déductible sur immos": "TVA déductible sur immobilisations",
    "44571 — TVA collectée": "TVA collectée",
    "44551 — TVA à décaisser": "Solde TVA à payer",
    "44567 — Crédit de TVA": "Crédit de TVA reportable",
    "44563 — TVA intracommunautaire": "TVA intracommunautaire déductible",
}

# ─────────────────────────────────────────────
# CALCULS
# ─────────────────────────────────────────────
def _calculer_tva(data: dict) -> dict:
    """Calcule TVA collectée, déductible, solde à partir des données saisies."""
    tva_collectee_detail = {}
    tva_collectee_total = 0.0
    for rubrique, info in RUBRIQUES_CA3.items():
        ht = data.get(rubrique, 0.0) or 0.0
        tva = ht * info["taux"]
        if ht != 0:
            tva_collectee_detail[rubrique] = {"base_ht": ht, "taux": info["taux"], "tva": tva}
        tva_collectee_total += tva

    tva_ded_abs = data.get("TVA déductible sur achats (ABS)", 0.0) or 0.0
    tva_ded_immo = data.get("TVA déductible sur immobilisations", 0.0) or 0.0
    tva_ded_intra = data.get("TVA intracommunautaire déductible", 0.0) or 0.0
    tva_deductible_total = tva_ded_abs + tva_ded_immo + tva_ded_intra
    credit_reporte = data.get("Crédit de TVA période précédente", 0.0) or 0.0
    solde = tva_collectee_total - tva_deductible_total - credit_reporte

    return {
        "tva_collectee_detail": tva_collectee_detail,
        "tva_collectee": tva_collectee_total,
        "tva_ded_abs": tva_ded_abs,
        "tva_ded_immo": tva_ded_immo,
        "tva_ded_intra": tva_ded_intra,
        "tva_deductible": tva_deductible_total,
        "credit_reporte": credit_reporte,
        "solde": solde,
        "a_payer": max(solde, 0),
        "credit_genere": max(-solde, 0),
    }


def _calculer_tva_depuis_balance_extrait(extrait: dict) -> dict:
    """Calcule le solde TVA à partir des montants extraits directement de la balance."""
    tva_collectee  = extrait.get("TVA collectée (44571)", 0) or 0.0
    tva_ded_abs    = extrait.get("TVA déductible ABS (44566)", 0) or 0.0
    tva_ded_immo   = extrait.get("TVA déductible immos (44562)", 0) or 0.0
    tva_ded_intra  = extrait.get("TVA intracommunautaire (44563)", 0) or 0.0
    credit_reporte = extrait.get("Crédit TVA (44567)", 0) or 0.0

    tva_deductible = tva_ded_abs + tva_ded_immo + tva_ded_intra
    solde = tva_collectee - tva_deductible - credit_reporte

    return {
        "tva_collectee_detail": {},
        "tva_collectee": tva_collectee,
        "tva_ded_abs": tva_ded_abs,
        "tva_ded_immo": tva_ded_immo,
        "tva_ded_intra": tva_ded_intra,
        "tva_deductible": tva_deductible,
        "credit_reporte": credit_reporte,
        "solde": solde,
        "a_payer": max(solde, 0),
        "credit_genere": max(-solde, 0),
    }


def _extraire_tva_depuis_balance(df: pd.DataFrame) -> dict:
    """Extrait les montants TVA depuis une balance PCG."""
    df = df.copy()

    def _col(df, names):
        for n in names:
            if n in df.columns:
                return n
        return None

    col_num  = _col(df, ['CompteNum', 'Compte', 'compte', 'NumCompte'])
    col_deb  = _col(df, ['Debit', 'debit', 'SoldeDebit', 'Mouvement_Debit'])
    col_cred = _col(df, ['Credit', 'credit', 'SoldeCredit', 'Mouvement_Credit'])

    if not col_num:
        return {}

    df['_num']  = df[col_num].astype(str).str.strip()
    df['_deb']  = pd.to_numeric(df[col_deb].astype(str).str.replace(',', '.').str.replace(' ', ''),
                                 errors='coerce').fillna(0) if col_deb else 0
    df['_cred'] = pd.to_numeric(df[col_cred].astype(str).str.replace(',', '.').str.replace(' ', ''),
                                 errors='coerce').fillna(0) if col_cred else 0
    df['_solde'] = df['_deb'] - df['_cred']

    def _somme(prefixes):
        mask = df['_num'].str.startswith(tuple(prefixes), na=False)
        return abs(df.loc[mask, '_solde'].sum())

    return {
        "TVA collectée (44571)":          _somme(['44571']),
        "TVA déductible ABS (44566)":     _somme(['44566']),
        "TVA déductible immos (44562)":   _somme(['44562']),
        "TVA intracommunautaire (44563)": _somme(['44563']),
        "Crédit TVA (44567)":             _somme(['44567']),
        "TVA à décaisser (44551)":        _somme(['44551']),
        "Chiffre d'affaires HT (70x)":    _somme(['70']),
        "Achats HT (60x)":                _somme(['60', '61', '62']),
    }


def _verifier_coherence(res: dict) -> list:
    """Retourne des alertes de cohérence."""
    alertes = []
    col = res.get("tva_collectee", 0)
    ded = res.get("tva_deductible", 0)
    solde = res.get("solde", 0)
    if col == 0 and ded > 0:
        alertes.append(("warning", "TVA collectée nulle mais déductible > 0 — vérifiez les comptes 44571"))
    if ded > col * 1.5 and col > 0:
        alertes.append(("warning", f"TVA déductible ({ded:,.0f} €) très supérieure à la TVA collectée ({col:,.0f} €) — crédit structurel, vérifiez"))
    if solde > 50000:
        alertes.append(("info", f"TVA à décaisser importante ({solde:,.0f} €) — pensez à la provision en comptabilité"))
    if res.get("credit_genere", 0) > 0:
        alertes.append(("success", f"Crédit de TVA de {res['credit_genere']:,.0f} € — remboursement possible si > 760 € (régime réel normal)"))
    return alertes

# ─────────────────────────────────────────────
# EXPORT EXCEL
# ─────────────────────────────────────────────
def _export_excel_tva(res: dict, data: dict, periode: str, entreprise: str) -> bytes:
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        rows_decl = []
        for rubrique, detail in res["tva_collectee_detail"].items():
            rows_decl.append({
                "Rubrique": rubrique,
                "Base HT (€)": detail["base_ht"],
                "Taux": f"{detail['taux']*100:.1f}%",
                "TVA collectée (€)": detail["tva"],
            })
        rows_decl.append({"Rubrique": "TOTAL TVA COLLECTÉE", "Base HT (€)": "", "Taux": "", "TVA collectée (€)": res["tva_collectee"]})
        rows_decl.append({"Rubrique": "TVA déductible ABS",  "Base HT (€)": "", "Taux": "", "TVA collectée (€)": -res["tva_ded_abs"]})
        rows_decl.append({"Rubrique": "TVA déductible immos","Base HT (€)": "", "Taux": "", "TVA collectée (€)": -res["tva_ded_immo"]})
        rows_decl.append({"Rubrique": "Crédit TVA reporté",  "Base HT (€)": "", "Taux": "", "TVA collectée (€)": -res["credit_reporte"]})
        rows_decl.append({"Rubrique": "SOLDE TVA",           "Base HT (€)": "", "Taux": "", "TVA collectée (€)": res["solde"]})
        pd.DataFrame(rows_decl).to_excel(writer, sheet_name="Déclaration TVA", index=False)

        ws = writer.sheets["Déclaration TVA"]
        from openpyxl.styles import PatternFill, Font
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E79")
            cell.font = Font(color="FFFFFF", bold=True)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 35

        rows_pcg = [{"Compte PCG": k, "Nature": v} for k, v in COMPTES_TVA_PCG.items()]
        pd.DataFrame(rows_pcg).to_excel(writer, sheet_name="Comptes PCG", index=False)

        pd.DataFrame([
            {"Champ": "Entreprise", "Valeur": entreprise},
            {"Champ": "Période",    "Valeur": periode},
            {"Champ": "Généré le",  "Valeur": datetime.now().strftime("%d/%m/%Y %H:%M")},
        ]).to_excel(writer, sheet_name="Info", index=False)
    return buf.getvalue()

# ─────────────────────────────────────────────
# GRAPHIQUE
# ─────────────────────────────────────────────
def _chart_tva(res: dict) -> go.Figure:
    labels = ["TVA Collectée", "TVA Déductible"]
    values = [res["tva_collectee"], res["tva_deductible"]]
    colors_bar = ["#e74c3c", "#27ae60"]
    fig = go.Figure(go.Bar(
        x=labels, y=values,
        marker_color=colors_bar,
        text=[f"{v:,.0f} €" for v in values],
        textposition="outside",
    ))
    solde = res["solde"]
    fig.add_hline(y=0, line_dash="dash", line_color="grey",
                  annotation_text=f"Solde : {solde:+,.0f} €",
                  annotation_position="bottom right")
    fig.update_layout(title="TVA Collectée vs Déductible", yaxis_title="Montant (€)", height=350)
    return fig

# ─────────────────────────────────────────────
# GÉNÉRATION PDF CA3 / CA12
# ─────────────────────────────────────────────
def _generer_pdf_ca3_ca12(res: dict, data: dict, periode: str, entreprise: str,
                           regime: str, siret: str = "", adresse: str = "") -> bytes:
    """Génère le formulaire CA3 ou CA12 en PDF au format DGFiP."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=2*cm, rightMargin=2*cm)

    bleu      = colors.HexColor('#003189')
    bleu_clair = colors.HexColor('#E8EEF7')
    rouge     = colors.HexColor('#CC0000')
    vert      = colors.HexColor('#006633')
    gris      = colors.HexColor('#F5F5F5')

    is_ca3    = "CA3" in regime
    type_decl = "CA3" if is_ca3 else "CA12"
    elements  = []

    def _p(text, size=8.5, bold=False, color=colors.black, align=TA_LEFT):
        font = 'Helvetica-Bold' if bold else 'Helvetica'
        return Paragraph(text, ParagraphStyle('s', fontSize=size, fontName=font,
                                               textColor=color, alignment=align))

    def section_header(titre):
        t = Table([[_p(titre, 9, bold=True, color=colors.white)]], colWidths=[18*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), bleu),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ]))
        return t

    def ligne(num, libelle, valeur, gras=False):
        bg = bleu_clair if gras else colors.white
        row = [
            _p(f"<b>{num}</b>", 7.5, color=colors.grey, align=TA_CENTER),
            _p(f"<b>{libelle}</b>" if gras else libelle, 8.5, bold=gras),
            _p(f"<b>{valeur}</b>" if gras else valeur, 8.5, bold=gras, align=TA_RIGHT),
        ]
        t = Table([row], colWidths=[1.2*cm, 12.8*cm, 4*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), bg),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
            ('LEFTPADDING', (1, 0), (1, 0), 4),
            ('LINEBELOW', (0, 0), (-1, -1), 0.3, colors.lightgrey),
            ('BOX', (2, 0), (2, 0), 0.5, colors.grey),
        ]))
        return t

    # ── EN-TÊTE ──
    hdr = Table([[
        _p("<b>DIRECTION GÉNÉRALE DES FINANCES PUBLIQUES</b>", 8, bold=True, color=bleu),
        _p(f"<b>DÉCLARATION DE TAXE SUR LA VALEUR AJOUTÉE<br/>Formulaire {type_decl}</b>",
           12, bold=True, color=bleu, align=TA_CENTER),
        _p(f"Période : <b>{periode}</b>", 9, align=TA_RIGHT),
    ]], colWidths=[5*cm, 9*cm, 4*cm])
    hdr.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), bleu_clair),
        ('BOX', (0, 0), (-1, -1), 1.5, bleu),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(hdr)
    elements.append(Spacer(1, 0.3*cm))

    # ── IDENTIFICATION ──
    elements.append(section_header("IDENTIFICATION DE L'ENTREPRISE"))
    elements.append(Spacer(1, 0.1*cm))
    id_t = Table([
        [_p("Dénomination :", 8.5, bold=True), _p(entreprise, 8.5),
         _p("SIRET :", 8.5, bold=True),         _p(siret or "________________", 8.5)],
        [_p("Adresse :", 8.5, bold=True),        _p(adresse or "________________", 8.5),
         _p("Régime :", 8.5, bold=True),          _p(type_decl, 8.5)],
        [_p("Période :", 8.5, bold=True),         _p(periode, 8.5),
         _p("Généré le :", 8.5, bold=True),       _p(datetime.now().strftime("%d/%m/%Y"), 8.5)],
    ], colWidths=[3*cm, 7*cm, 2.5*cm, 5.5*cm])
    id_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), gris),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(id_t)
    elements.append(Spacer(1, 0.3*cm))

    # ── I. BASES HT ──
    elements.append(section_header("I. OPÉRATIONS RÉALISÉES — BASES HORS TAXE"))
    elements.append(Spacer(1, 0.05*cm))

    def get_base(k):
        if k in res.get("tva_collectee_detail", {}):
            return res["tva_collectee_detail"][k]["base_ht"]
        return data.get(k, 0) or 0

    for num, lib, key in [
        ("A1", "Opérations imposables à 20% — taux normal",           "Ventes soumises à 20%"),
        ("A2", "Opérations imposables à 10% — taux intermédiaire",    "Ventes soumises à 10%"),
        ("A3", "Opérations imposables à 5,5% — taux réduit",          "Ventes soumises à 5,5%"),
        ("A4", "Opérations imposables à 2,1% — taux super-réduit",    "Ventes soumises à 2,1%"),
        ("B1", "Acquisitions intracommunautaires (20%)",               "Acquisitions intracommunautaires (20%)"),
        ("B2", "Autoliquidation — Achats à 20%",                      "Autoliquidation achats (20%)"),
        ("E1", "Opérations exonérées / Exportations hors UE",         "Ventes exonérées / exports"),
    ]:
        val = get_base(key)
        elements.append(ligne(num, lib, f"{val:,.2f} €" if val else "—"))
    elements.append(Spacer(1, 0.3*cm))

    # ── II. TVA BRUTE ──
    elements.append(section_header("II. TVA BRUTE"))
    elements.append(Spacer(1, 0.05*cm))

    def get_tva(k):
        if k in res.get("tva_collectee_detail", {}):
            return res["tva_collectee_detail"][k]["tva"]
        return 0

    tva20 = (get_tva("Ventes soumises à 20%")
             + get_tva("Acquisitions intracommunautaires (20%)")
             + get_tva("Autoliquidation achats (20%)"))
    tva10 = get_tva("Ventes soumises à 10%")
    tva55 = get_tva("Ventes soumises à 5,5%")
    tva21 = get_tva("Ventes soumises à 2,1%")
    if not res.get("tva_collectee_detail"):
        tva20 = res["tva_collectee"]

    for num, lib, val in [
        ("08", "TVA au taux normal 20%",           tva20),
        ("09", "TVA au taux intermédiaire 10%",    tva10),
        ("10", "TVA au taux réduit 5,5%",          tva55),
        ("11", "TVA au taux super-réduit 2,1%",    tva21),
    ]:
        elements.append(ligne(num, lib, f"{val:,.2f} €" if val else "—"))
    elements.append(ligne("16", "TOTAL TVA BRUTE", f"{res['tva_collectee']:,.2f} €", gras=True))
    elements.append(Spacer(1, 0.3*cm))

    # ── III. TVA DÉDUCTIBLE ──
    elements.append(section_header("III. TVA DÉDUCTIBLE"))
    elements.append(Spacer(1, 0.05*cm))
    for num, lib, val in [
        ("19", "TVA déductible sur immobilisations — compte 44562",          res["tva_ded_immo"]),
        ("20", "TVA déductible sur autres biens et services — compte 44566", res["tva_ded_abs"]),
        ("21", "TVA intracommunautaire déductible — compte 44563",           res["tva_ded_intra"]),
        ("22", "Crédit de TVA période précédente — compte 44567",            res["credit_reporte"]),
    ]:
        elements.append(ligne(num, lib, f"{val:,.2f} €" if val else "—"))
    elements.append(ligne("23", "TOTAL TVA DÉDUCTIBLE", f"{res['tva_deductible']:,.2f} €", gras=True))
    elements.append(Spacer(1, 0.3*cm))

    # ── IV. RÉSULTAT ──
    elements.append(section_header("IV. RÉSULTAT DE LA DÉCLARATION"))
    elements.append(Spacer(1, 0.1*cm))
    if res["a_payer"] > 0:
        res_color, res_bg, res_num = rouge, colors.HexColor('#FFF0F0'), "25"
        res_lib = "TVA A DECAISSER — Compte 44551"
        res_val = f"{res['a_payer']:,.2f} €"
    else:
        res_color, res_bg, res_num = vert, colors.HexColor('#F0FFF0'), "26"
        res_lib = "CREDIT DE TVA — Compte 44567"
        res_val = f"{res['credit_genere']:,.2f} €"
    res_t = Table([[
        _p(f"<b>{res_num}</b>", 8, color=colors.grey, align=TA_CENTER),
        _p(f"<b>{res_lib}</b>", 11, bold=True, color=res_color),
        _p(f"<b>{res_val}</b>", 13, bold=True, color=res_color, align=TA_RIGHT),
    ]], colWidths=[1.2*cm, 12.8*cm, 4*cm])
    res_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), res_bg),
        ('BOX', (0, 0), (-1, -1), 1.5, res_color),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    elements.append(res_t)

    if res.get("credit_genere", 0) >= 760:
        elements.append(Spacer(1, 0.2*cm))
        elements.append(_p(
            "Info : Credit superieur a 760 EUR — Remboursement possible sur demande (CA3 reel normal)",
            7.5, color=vert))
    elements.append(Spacer(1, 0.3*cm))

    # ── V. ÉCRITURE COMPTABLE ──
    elements.append(section_header("V. ECRITURE COMPTABLE A PASSER"))
    elements.append(Spacer(1, 0.1*cm))
    if res["a_payer"] > 0:
        ecr = (f"Debit  44551 TVA a decaisser     {res['a_payer']:>12,.2f} EUR\n"
               f"  Credit  512  Banque             {res['a_payer']:>12,.2f} EUR\n"
               f"  Libelle : Reglement TVA {periode}")
    else:
        ecr = (f"Debit  44567 Credit de TVA        {res['credit_genere']:>12,.2f} EUR\n"
               f"  Credit 44551 TVA a decaisser    {res['credit_genere']:>12,.2f} EUR\n"
               f"  Libelle : Report credit TVA {periode}")
    ecr_t = Table([[_p(f"<font name='Courier' size='8'>{ecr}</font>", 8)]],
                  colWidths=[18*cm])
    ecr_t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), gris),
        ('BOX', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(ecr_t)
    elements.append(Spacer(1, 0.4*cm))

    # ── PIED DE PAGE ──
    footer = Table([[
        _p("Genere par <b>SMD Global Consulting LLC</b> — Superviseur IA Comptable", 7, color=colors.grey),
        _p("ATTENTION : Document d aide. La declaration officielle doit etre deposee sur impots.gouv.fr",
           7, color=colors.grey, align=TA_RIGHT),
    ]], colWidths=[9*cm, 9*cm])
    footer.setStyle(TableStyle([
        ('LINEABOVE', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(footer)

    doc.build(elements)
    return buf.getvalue()

# ─────────────────────────────────────────────
# PAGE PRINCIPALE
# ─────────────────────────────────────────────
def page_tva():
    st.title("🧾 Aide Déclaration TVA — CA3 / CA12")
    st.markdown("**Calcul et vérification TVA France** — Régimes réel normal (CA3) et simplifié (CA12)")
    st.caption("✨ Pour Cabinets, DAF et Dirigeants - DGFiP / PCG France")

    # ── Paramètres ──
    col1, col2, col3 = st.columns(3)
    with col1:
        entreprise = st.text_input("🏢 Entreprise", value="Entreprise")
    with col2:
        regime = st.selectbox("📋 Régime TVA", [
            "CA3 — Réel normal (mensuel/trimestriel)",
            "CA12 — Réel simplifié (annuel)"
        ])
    with col3:
        periode = st.text_input("📅 Période", value=f"{datetime.now().strftime('%m/%Y')}")

    col4, col5 = st.columns(2)
    with col4:
        siret = st.text_input("🆔 SIRET", placeholder="Ex: 123 456 789 00012")
    with col5:
        adresse = st.text_input("📍 Adresse", placeholder="Ex: 12 rue de la Paix, 75001 Paris")

    st.divider()

    # ── Tabs ──
    tab1, tab2, tab3 = st.tabs(["📝 Saisie manuelle", "📁 Import balance", "📖 Guide taux & comptes"])

    # ══════════════════════════════════════════
    # TAB 1 — Saisie manuelle
    # ══════════════════════════════════════════
    with tab1:
        st.markdown("### 💰 Bases HT soumises à TVA (ventes / opérations imposables)")
        data = {}
        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("**Ventes / Opérations taxables**")
            data["Ventes soumises à 20%"]    = st.number_input("Ventes 20% — Base HT (€)", min_value=0.0, step=100.0, format="%.2f")
            data["Ventes soumises à 10%"]    = st.number_input("Ventes 10% — Base HT (€)", min_value=0.0, step=100.0, format="%.2f")
            data["Ventes soumises à 5,5%"]   = st.number_input("Ventes 5,5% — Base HT (€)", min_value=0.0, step=100.0, format="%.2f")
            data["Ventes soumises à 2,1%"]   = st.number_input("Ventes 2,1% — Base HT (€)", min_value=0.0, step=100.0, format="%.2f")
            data["Ventes exonérées / exports"] = st.number_input("Ventes exonérées / exports (€ HT)", min_value=0.0, step=100.0, format="%.2f")
        with col_b:
            st.markdown("**Opérations particulières**")
            data["Acquisitions intracommunautaires (20%)"] = st.number_input("Acquisitions intracom. 20% — Base HT (€)", min_value=0.0, step=100.0, format="%.2f")
            data["Autoliquidation achats (20%)"]           = st.number_input("Autoliquidation achats 20% — Base HT (€)", min_value=0.0, step=100.0, format="%.2f")

        st.divider()
        st.markdown("### 🔻 TVA Déductible")
        col_d1, col_d2, col_d3 = st.columns(3)
        with col_d1:
            data["TVA déductible sur achats (ABS)"]      = st.number_input("TVA ded. ABS — 44566 (€)", min_value=0.0, step=10.0, format="%.2f")
        with col_d2:
            data["TVA déductible sur immobilisations"]   = st.number_input("TVA ded. Immos — 44562 (€)", min_value=0.0, step=10.0, format="%.2f")
        with col_d3:
            data["TVA intracommunautaire déductible"]    = st.number_input("TVA intracom. ded. — 44563 (€)", min_value=0.0, step=10.0, format="%.2f")
        data["Crédit de TVA période précédente"] = st.number_input("Crédit TVA période précédente — 44567 (€)", min_value=0.0, step=10.0, format="%.2f")

        st.divider()
        if st.button("🧮 Calculer la déclaration TVA", type="primary", use_container_width=True):
            res = _calculer_tva(data)
            _afficher_resultats(res, data, periode, entreprise, regime, siret, adresse)

    # ══════════════════════════════════════════
    # TAB 2 — Import balance
    # ══════════════════════════════════════════
    with tab2:
        st.markdown("### 📁 Import balance comptable")
        st.info("La balance est analysée pour extraire automatiquement les comptes TVA (445xx) et CA (70x).")
        uploaded = st.file_uploader("Balance CSV / Excel", type=["csv", "xlsx"], key="tva_balance")
        if uploaded:
            try:
                from utils.intelligent_parser import parser_balance_intelligent
                with st.spinner("Analyse de la balance..."):
                    df, info = parser_balance_intelligent(uploaded)
                st.success(f"✅ {len(df):,} comptes chargés — {info.get('format_detecte', 'format détecté')}")
                extrait = _extraire_tva_depuis_balance(df)
                if extrait:
                    st.markdown("### 📊 Comptes TVA extraits de la balance")
                    df_ext = pd.DataFrame([
                        {"Compte / Rubrique": k, "Montant (€)": f"{v:,.2f}"}
                        for k, v in extrait.items() if v != 0
                    ])
                    st.dataframe(df_ext, use_container_width=True, hide_index=True)

                    st.divider()
                    st.markdown("#### ✏ Vérifier / Ajuster les montants avant déclaration")
                    col1, col2 = st.columns(2)
                    with col1:
                        tva_col  = st.number_input("TVA Collectée — 44571 (€)",
                                                   value=float(extrait.get("TVA collectée (44571)", 0)),
                                                   min_value=0.0, step=10.0, format="%.2f", key="imp_tva_col")
                        tva_abs  = st.number_input("TVA Déductible ABS — 44566 (€)",
                                                   value=float(extrait.get("TVA déductible ABS (44566)", 0)),
                                                   min_value=0.0, step=10.0, format="%.2f", key="imp_tva_abs")
                        tva_immo = st.number_input("TVA Déductible Immos — 44562 (€)",
                                                   value=float(extrait.get("TVA déductible immos (44562)", 0)),
                                                   min_value=0.0, step=10.0, format="%.2f", key="imp_tva_immo")
                    with col2:
                        tva_intra = st.number_input("TVA Intracommunautaire — 44563 (€)",
                                                    value=float(extrait.get("TVA intracommunautaire (44563)", 0)),
                                                    min_value=0.0, step=10.0, format="%.2f", key="imp_tva_intra")
                        credit_r  = st.number_input("Crédit TVA reporté — 44567 (€)",
                                                    value=float(extrait.get("Crédit TVA (44567)", 0)),
                                                    min_value=0.0, step=10.0, format="%.2f", key="imp_credit")

                    st.divider()
                    if st.button("🧮 Générer la déclaration CA3/CA12", type="primary",
                                 use_container_width=True, key="btn_import_decl"):
                        extrait_ajuste = {
                            "TVA collectée (44571)":            tva_col,
                            "TVA déductible ABS (44566)":       tva_abs,
                            "TVA déductible immos (44562)":     tva_immo,
                            "TVA intracommunautaire (44563)":   tva_intra,
                            "Crédit TVA (44567)":               credit_r,
                        }
                        res = _calculer_tva_depuis_balance_extrait(extrait_ajuste)
                        _afficher_resultats(res, extrait_ajuste, periode, entreprise, regime, siret, adresse)
                else:
                    st.warning("Aucun compte TVA (445xx) détecté dans la balance.")
            except Exception as e:
                st.error(f"Erreur : {e}")
                import traceback
                with st.expander("Détails"):
                    st.code(traceback.format_exc())

    # ══════════════════════════════════════════
    # TAB 3 — Guide
    # ══════════════════════════════════════════
    with tab3:
        st.markdown("### 📖 Taux de TVA applicables en France")
        df_taux = pd.DataFrame([
            {"Taux": "20%",   "Catégorie": "Taux normal",       "Exemples": "Biens et services courants, honoraires, conseil"},
            {"Taux": "10%",   "Catégorie": "Taux intermédiaire","Exemples": "Restauration, transport, travaux logement, médicaments remboursables"},
            {"Taux": "5,5%",  "Catégorie": "Taux réduit",       "Exemples": "Alimentation, livres, abonnements énergie, équipement PMR"},
            {"Taux": "2,1%",  "Catégorie": "Taux super-réduit", "Exemples": "Presse, médicaments remboursés SS, spectacles vivants (100 premières représentations)"},
            {"Taux": "0%",    "Catégorie": "Exonéré",           "Exemples": "Exports hors UE, intracommunautaire, activités médicales, enseignement, assurance"},
        ])
        st.dataframe(df_taux, use_container_width=True, hide_index=True)
        st.divider()

        st.markdown("### 🏦 Comptes PCG — TVA")
        df_pcg = pd.DataFrame([{"Compte": k, "Nature": v} for k, v in COMPTES_TVA_PCG.items()])
        st.dataframe(df_pcg, use_container_width=True, hide_index=True)
        st.divider()

        st.markdown("### 📋 CA3 vs CA12 — Comparatif régimes")
        df_regime = pd.DataFrame([
            {"Critère": "Fréquence",             "CA3 (Réel normal)": "Mensuelle ou trimestrielle",                    "CA12 (Réel simplifié)": "Annuelle (mai N+1)"},
            {"Critère": "CA seuil (BIC/BNC)",    "CA3 (Réel normal)": "> 840 000 € (négoce) / 254 000 € (services)", "CA12 (Réel simplifié)": "< 840 000 € (négoce) / 254 000 € (services)"},
            {"Critère": "Acomptes",              "CA3 (Réel normal)": "Aucun — déclaration mensuelle",                 "CA12 (Réel simplifié)": "2 acomptes (55% en juil., 40% en déc.)"},
            {"Critère": "Crédit TVA",            "CA3 (Réel normal)": "Remboursable dès 760 €",                       "CA12 (Réel simplifié)": "Sur demande ou imputation"},
            {"Critère": "Comptabilité",          "CA3 (Réel normal)": "Obligatoirement complète",                     "CA12 (Réel simplifié)": "Simplifiée possible"},
        ])
        st.dataframe(df_regime, use_container_width=True, hide_index=True)
        st.divider()

        st.markdown("### ⚡ Autoliquidation — Cas principaux")
        df_auto = pd.DataFrame([
            {"Opération": "Acquisitions intracommunautaires (AIC)", "Mécanisme": "Autoliquidation — acheteur déclare et déduit simultanément"},
            {"Opération": "Sous-traitance BTP",                     "Mécanisme": "Donneur d'ordre autoliquide la TVA du sous-traitant"},
            {"Opération": "Services étrangers (art. 283-1)",        "Mécanisme": "Preneur français autoliquide si prestataire non établi en France"},
            {"Opération": "Livraisons intracommunautaires (LIC)",   "Mécanisme": "Exonéré côté vendeur — TVA du pays acheteur"},
        ])
        st.dataframe(df_auto, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────
# AFFICHAGE RÉSULTATS (partagé saisie/import)
# ─────────────────────────────────────────────
def _afficher_resultats(res: dict, data: dict, periode: str, entreprise: str,
                         regime: str, siret: str = "", adresse: str = ""):
    st.divider()
    st.markdown(f"## 📋 Déclaration TVA — {entreprise} — {periode}")
    st.caption(regime)

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("💰 TVA Collectée",  f"{res['tva_collectee']:,.0f} €")
    col2.metric("🔻 TVA Déductible", f"{res['tva_deductible']:,.0f} €")
    if res['a_payer'] > 0:
        col3.metric("💸 TVA à Décaisser", f"{res['a_payer']:,.0f} €", delta="À payer",  delta_color="inverse")
        col4.metric("✅ Crédit TVA",       "0 €")
    else:
        col3.metric("💸 TVA à Décaisser", "0 €")
        col4.metric("✅ Crédit TVA",       f"{res['credit_genere']:,.0f} €", delta="Crédit", delta_color="normal")

    st.divider()

    if res["tva_collectee_detail"]:
        st.markdown("### 🔺 Détail TVA Collectée")
        rows = []
        for rubrique, detail in res["tva_collectee_detail"].items():
            rows.append({
                "Rubrique":    rubrique,
                "Base HT (€)": f"{detail['base_ht']:,.2f}",
                "Taux":        f"{detail['taux']*100:.1f}%",
                "TVA (€)":     f"{detail['tva']:,.2f}",
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    st.markdown("### 🔻 Détail TVA Déductible")
    rows_ded = []
    if res["tva_ded_abs"]    > 0: rows_ded.append({"Compte": "44566", "Nature": "TVA déductible sur ABS",             "Montant (€)": f"{res['tva_ded_abs']:,.2f}"})
    if res["tva_ded_immo"]   > 0: rows_ded.append({"Compte": "44562", "Nature": "TVA déductible sur immos",           "Montant (€)": f"{res['tva_ded_immo']:,.2f}"})
    if res["tva_ded_intra"]  > 0: rows_ded.append({"Compte": "44563", "Nature": "TVA intracom. déductible",           "Montant (€)": f"{res['tva_ded_intra']:,.2f}"})
    if res["credit_reporte"] > 0: rows_ded.append({"Compte": "44567", "Nature": "Crédit TVA période précédente",      "Montant (€)": f"{res['credit_reporte']:,.2f}"})
    if rows_ded:
        st.dataframe(pd.DataFrame(rows_ded), use_container_width=True, hide_index=True)
    else:
        st.info("Aucune TVA déductible saisie.")

    st.markdown("### 🏁 Synthèse")
    cols_s = st.columns(3)
    for i, (k, v) in enumerate({
        "TVA Collectée":      res["tva_collectee"],
        "TVA Déductible":     res["tva_deductible"],
        "Solde (+ = à payer)": res["solde"],
    }.items()):
        cols_s[i].metric(k, f"{v:,.2f} €")

    st.plotly_chart(_chart_tva(res), use_container_width=True)

    alertes = _verifier_coherence(res)
    if alertes:
        st.divider()
        st.markdown("### ⚠ Contrôles de cohérence")
        for typ, msg in alertes:
            if typ == "warning":   st.warning(msg)
            elif typ == "success": st.success(msg)
            else:                  st.info(msg)

    st.divider()
    st.markdown("### 📝 Écriture comptable à passer (PCG)")
    if res["a_payer"] > 0:
        st.code(f"""
Débit  44551 — TVA à décaisser      {res['a_payer']:>12,.2f} €
  Crédit 512 — Banque                        {res['a_payer']:>12,.2f} €
  → Règlement TVA {periode}
""", language="text")
    else:
        st.code(f"""
Débit  44567 — Crédit de TVA        {res['credit_genere']:>12,.2f} €
  Crédit 44551 — TVA à décaisser             {res['credit_genere']:>12,.2f} €
  → Report crédit TVA {periode}
""", language="text")

    # ── Export Excel ──
    st.divider()
    try:
        excel = _export_excel_tva(res, data, periode, entreprise)
        st.download_button(
            label="📥 Télécharger Excel Déclaration TVA",
            data=excel,
            file_name=f"TVA_{entreprise}_{periode.replace('/', '-')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    except Exception as e:
        st.warning(f"Export Excel non disponible : {e}")

    # ── Export PDF CA3/CA12 ──
    try:
        type_decl = "CA3" if "CA3" in regime else "CA12"
        pdf_bytes = _generer_pdf_ca3_ca12(res, data, periode, entreprise, regime, siret, adresse)
        st.download_button(
            label=f"📄 Télécharger le formulaire {type_decl} (PDF DGFiP)",
            data=pdf_bytes,
            file_name=f"{type_decl}_{entreprise}_{periode.replace('/', '-')}.pdf",
            mime="application/pdf",
            use_container_width=True,
            type="primary",
        )
    except Exception as e:
        st.warning(f"Export PDF non disponible : {e}")
